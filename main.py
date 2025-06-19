from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import uuid
import tempfile
import io

from Room_dimension.room_dimension import (
    extract_room_boundaries,
    extract_block_text_info,
    clean_text_value,
    check_text_within_room
)
from Room_dimension.dwg_logic import save_copy_with_changes

from db import add_history, get_all_history, clear_all_history, init_db, delete_entry, get_file_content


app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")


@app.get("/", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/history", response_class=HTMLResponse)
async def view_history(request: Request):
    records = get_all_history()
    return templates.TemplateResponse("history.html", {"request": request, "records": records})


@app.post("/login", response_class=HTMLResponse)
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == "admin" and password == "admin":
        return RedirectResponse(url="/home", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": "Login failed"})


@app.get("/home", response_class=HTMLResponse)
async def home_page(request: Request):
    history_data = get_all_history()
    return templates.TemplateResponse("home.html", {"request": request, "history": history_data})


@app.get("/clear-history")
async def clear_history():
    clear_all_history()
    return RedirectResponse("/home", status_code=302)

@app.get("/delete-entry/{entry_id}")
async def delete_entry_route(entry_id: int):
    delete_entry(entry_id)
    return RedirectResponse(url="/history", status_code=302)


@app.get("/self-check", response_class=HTMLResponse)
async def self_check(request: Request):
    return templates.TemplateResponse("result.html", {
        "request": request,
        "check_type": "Self Check",
        "filename": None,
        "rooms": None,
        "submitted_rooms": None,
        "matches": None,
        "download_link": None,
        "excel_link": None,
        "client_file_link": None
    })

from fastapi import UploadFile, File, Form

from Referece_check.reference import extract_room_dimensions, extract_doors, compare_values, visualize_mismatches

@app.get("/reference-check", response_class=HTMLResponse)
async def reference_check_page(request: Request):
    return templates.TemplateResponse("reference.html", {
        "request": request,
        "room_mismatches": None,
        "door_mismatches": None,
        "download_link": None,
        "excel_link": None,
        "client_file_link": None
    })


@app.post("/reference-check", response_class=HTMLResponse)
async def reference_check_upload(
    request: Request,
    ref_file: UploadFile = File(...),
    client_file: UploadFile = File(...)
):
    try:
        # Check valid file types
        if not ref_file.filename.endswith(".dxf") or not client_file.filename.endswith(".dxf"):
            return templates.TemplateResponse("reference.html", {
                "request": request,
                "error": "Only .dxf files are supported.",
                "room_mismatches": None,
                "door_mismatches": None,
                "download_link": None
            })

        # Read file contents
        ref_content = await ref_file.read()
        client_content = await client_file.read()

        # Create temporary files for processing
        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as ref_temp:
            ref_temp.write(ref_content)
            ref_temp_path = ref_temp.name

        with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as client_temp:
            client_temp.write(client_content)
            client_temp_path = client_temp.name

        try:
            # DXF comparison
            import ezdxf
            ref_doc = ezdxf.readfile(ref_temp_path)
            client_doc = ezdxf.readfile(client_temp_path)

            ref_rooms = extract_room_dimensions(ref_doc)
            client_rooms = extract_room_dimensions(client_doc)
            room_mismatches = compare_values(ref_rooms, client_rooms)

            ref_doors = extract_doors(ref_doc)
            client_doors = extract_doors(client_doc)
            door_mismatches = compare_values(ref_doors, client_doors)

            # Save updated DXF with visualization
            with tempfile.NamedTemporaryFile(suffix="_visualized.dxf", delete=False) as updated_temp:
                updated_temp_path = updated_temp.name

            visualize_mismatches(client_temp_path, room_mismatches, door_mismatches, updated_temp_path)
            
            with open(updated_temp_path, 'rb') as f:
                updated_content = f.read()

            # Generate Excel report
            excel_data = pd.DataFrame({
                "Room Mismatches": room_mismatches,
                "Door Mismatches": door_mismatches
            })
            
            excel_buffer = io.BytesIO()
            excel_data.to_excel(excel_buffer, index=False)
            excel_content = excel_buffer.getvalue()

            # Store in database
            history_id = add_history(
                "Reference Check",
                ref_file.filename,
                client_file.filename,
                client_file.filename.replace('.dxf', '_visualized.dxf'),
                f"report_{uuid.uuid4().hex}.xlsx",
                ref_content,
                client_content,
                updated_content,
                excel_content
            )

            return templates.TemplateResponse("reference.html", {
                "request": request,
                "check_type": "Reference Check",
                "filename": client_file.filename,
                "rooms": None,
                "submitted_rooms": None,
                "matches": None,
                "room_mismatches": room_mismatches,
                "door_mismatches": door_mismatches,
                "download_link": f"/download/modified/{history_id}",
                "excel_link": f"/download/excel/{history_id}",
                "client_file_link": f"/download/client/{history_id}"
            })

        finally:
            # Clean up temporary files
            os.unlink(ref_temp_path)
            os.unlink(client_temp_path)
            if 'updated_temp_path' in locals():
                os.unlink(updated_temp_path)

    except Exception as e:
        return templates.TemplateResponse("reference.html", {
            "request": request,
            "error": f"Error: {str(e)}",
            "room_mismatches": None,
            "door_mismatches": None,
            "download_link": None
        })


def feet_inches_to_mm(feet: int, inches: int) -> float:
    return round((feet * 12 + inches) * 25.4, 2)


def match_user_rooms_to_dxf(submitted_rooms, extracted_rooms):
    matched = []
    unmatched = []

    for user_room in submitted_rooms:
        is_matched = False
        for dxf_room in extracted_rooms:
            room_name_matches = any(
                user_room["name"] in t["cleaned"].lower()
                for t in dxf_room["texts"]
            )
            area_match = abs(user_room["width_mm"] * user_room["height_mm"] - dxf_room["Area"]) <= 100000

            if room_name_matches and area_match:
                matched.append({
                    "user_room": user_room,
                    "matched_room": dxf_room
                })
                is_matched = True
                break

        if not is_matched:
            unmatched.append(user_room)

    return matched, unmatched


def mm_to_feet(mm):
    return round(mm / 304.8, 2)

def mm2_to_sqft(area_mm2):
    return round(area_mm2 / 92903.04, 2)

def export_matches_to_excel(matches, unmatched):
    matched_data = []
    unmatched_data = []

    for match in matches:
        user = match["user_room"]
        dxf = match["matched_room"]
        block_name = dxf.get("BlockName") or dxf.get("Block") or "N/A"

        matched_data.append({
            "Block": block_name,
            "Name of room": user["name"],
            "Length from dxf(feet)": mm_to_feet(dxf.get("Length", 0)),
            "breadth from dxf(feet)": mm_to_feet(dxf.get("Breadth", 0)),
            "Input length(feet)": mm_to_feet(user["width_mm"]),
            "Input breadth(feet)": mm_to_feet(user["height_mm"]),
            "Match": "YES",
            "Reason": "Match",
            "area": mm2_to_sqft(dxf.get("Area", 0))
        })

    for user in unmatched:
        unmatched_data.append({
            "Block": user.get("Block", "N/A"),
            "Name of room": user["name"],
            "Length from dxf(feet)": mm_to_feet(user.get("dxf_length", 0)) if user.get("dxf_length") else "",
            "breadth from dxf(feet)": mm_to_feet(user.get("dxf_breadth", 0)) if user.get("dxf_breadth") else "",
            "Input length(feet)": mm_to_feet(user["width_mm"]),
            "Input breadth(feet)": mm_to_feet(user["height_mm"]),
            "Match": "NO",
            "Reason": user.get("reason", "User Input not match"),
            "area": ""
        })

    # Combine into one DataFrame with separation
    rows = []
    rows.append({"Block": "Matched Room"})  # Section Header
    rows.extend(matched_data)
    rows.append({})  # Empty row
    rows.append({"Block": "Not Matched Room"})  # Section Header
    rows.extend(unmatched_data)

    df = pd.DataFrame(rows)

    # Create Excel in memory
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')
    
    # Apply colors
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        match_cell = row[6]  # Match column
        if match_cell.value == "YES":
            for cell in row:
                cell.fill = green_fill
        elif match_cell.value == "NO":
            for cell in row:
                cell.fill = red_fill

    # Save to buffer
    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    
    return final_buffer.getvalue()

@app.post("/upload-{check_type}", response_class=HTMLResponse)
async def upload_dxf(
    request: Request,
    check_type: str,
    file: UploadFile = File(...),
    room_count: int = Form(...),
):
    form = await request.form()
    check_type_display = check_type.replace("-", " ").title()

    if not file.filename.lower().endswith(".dxf"):
        return templates.TemplateResponse("result.html", {
            "request": request,
            "check_type": check_type_display,
            "error": "Only .dxf files are supported.",
            "filename": None,
            "rooms": None,
            "submitted_rooms": None,
            "matches": None,
            "download_link": None,
            "excel_link": None,
            "client_file_link": None
        })

    # Read file content
    client_content = await file.read()

    # Create temporary file for processing
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as temp_file:
        temp_file.write(client_content)
        temp_path = temp_file.name

    try:
        rooms = extract_room_boundaries(temp_path)
        texts = extract_block_text_info(temp_path)

        all_texts = []
        for blk_texts in texts.values():
            for text in blk_texts:
                text["original"] = text["Text"]
                text["cleaned"] = clean_text_value(text["Text"])
                all_texts.append(text)

        for room in rooms:
            room["texts"] = [t for t in all_texts if check_text_within_room(room, t)]

        submitted_rooms = []
        for i in range(1, room_count + 1):
            name = form.get(f"room_name_{i}", "").strip().lower()
            width_ft = int(form.get(f"width_feet_{i}", 0))
            width_in = int(form.get(f"width_inches_{i}", 0))
            height_ft = int(form.get(f"height_feet_{i}", 0))
            height_in = int(form.get(f"height_inches_{i}", 0))

            submitted_rooms.append({
                "name": name,
                "width_mm": feet_inches_to_mm(width_ft, width_in),
                "height_mm": feet_inches_to_mm(height_ft, height_in),
                "width_feet": width_ft,
                "width_inches": width_in,
                "height_feet": height_ft,
                "height_inches": height_in
            })

        matches, unmatched = match_user_rooms_to_dxf(submitted_rooms, rooms)

        # Generate Excel report in memory
        excel_content = export_matches_to_excel(matches, unmatched)

        # Create updated DXF file
        with tempfile.NamedTemporaryFile(suffix="_updated.dxf", delete=False) as updated_temp:
            updated_temp_path = updated_temp.name

        save_copy_with_changes(temp_path, updated_temp_path, rooms, texts)
        
        with open(updated_temp_path, 'rb') as f:
            updated_content = f.read()

        reference_file = "" if check_type == "self-check" else form.get("reference_filename", "")

        # Store in database
        history_id = add_history(
            check_type_display,
            reference_file,
            file.filename,
            f"updated_{file.filename}",
            "full_report.xlsx",
            b"",  # No reference file content for self-check
            client_content,
            updated_content,
            excel_content
        )

        return templates.TemplateResponse("result.html", {
            "request": request,
            "check_type": check_type_display,
            "filename": file.filename,
            "rooms": rooms,
            "submitted_rooms": submitted_rooms,
            "matches": matches,
            "download_link": f"/download/modified/{history_id}",
            "excel_link": f"/download/excel/{history_id}",
            "client_file_link": f"/download/client/{history_id}"
        })

    finally:
        # Clean up temporary files
        os.unlink(temp_path)
        if 'updated_temp_path' in locals():
            os.unlink(updated_temp_path)


@app.get("/download/reference/{history_id}")
async def download_reference_file(history_id: int):
    file_content, filename = get_file_content(history_id, "reference")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


@app.get("/download/client/{history_id}")
async def download_client_file(history_id: int):
    file_content, filename = get_file_content(history_id, "client")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


@app.get("/download/modified/{history_id}")
async def download_modified_file(history_id: int):
    file_content, filename = get_file_content(history_id, "modified")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


@app.get("/download/excel/{history_id}")
async def download_excel_file(history_id: int):
    file_content, filename = get_file_content(history_id, "excel")
    if file_content:
        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    return HTMLResponse("<h3>File not found</h3>", status_code=404)


# Legacy download endpoint for backward compatibility
@app.get("/download/{filename}")
async def download_file(filename: str):
    return HTMLResponse(f"<h3>File not found: {filename}</h3>", status_code=404)


init_db()